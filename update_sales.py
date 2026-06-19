# -*- coding: utf-8 -*-
"""
판매 통계 자동 업데이트

사용법:
    python update_sales.py

파일 자동 탐지 (판매통계 BD 폴더 기준):
    일별주문통계_커넥트킨록_*.xls   -> 무신사 커넥트킨록
    일별주문통계_에든버러클럽_*.xls  -> 무신사 에든버러클럽
    글로벌주문내역_*.xls             -> 글로벌
    29CM_커넥트킨록_*.xlsx           -> 29CM 커넥트킨록
    29CM_에든버러클럽_*.xlsx         -> 29CM 에든버러클럽
"""

import gc
import os
import pathlib
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from copy import copy
from datetime import date as _date

import openpyxl
import pandas as pd

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
_ONEDRIVE_DESKTOP = pathlib.Path.home() / "Library/CloudStorage/OneDrive-개인/바탕 화면"
SCRIPT_DIR  = str(_ONEDRIVE_DESKTOP / "판매통계 BD")
DESKTOP_DIR = str(_ONEDRIVE_DESKTOP)
STATS_FILE  = os.path.join(DESKTOP_DIR, "판매 통계.xlsx")
BACKUP_FILE = os.path.join(DESKTOP_DIR, "판매 통계 백업.xlsx")


# ── 파일 자동 탐지 ─────────────────────────────────────────────────────────────
def find_source_files():
    import unicodedata
    bd = SCRIPT_DIR
    try:
        entries = os.listdir(bd)
    except Exception:
        entries = []

    def latest(prefix, suffix):
        nfc_prefix = unicodedata.normalize('NFC', prefix)
        matches = sorted(
            e for e in entries
            if unicodedata.normalize('NFC', e).startswith(nfc_prefix)
            and unicodedata.normalize('NFC', e).endswith(suffix)
        )
        return os.path.join(bd, matches[-1]) if matches else None

    return {
        "mushinsa":  latest("일별주문통계_커넥트킨록_",  ".xls"),
        "mushinsa2": latest("일별주문통계_에든버러클럽_", ".xls"),
        "glob":      latest("글로벌주문내역_",           ".xls"),
        "cm1":       latest("29CM_커넥트킨록_",          ".xlsx"),
        "cm2":       latest("29CM_에든버러클럽_",         ".xlsx"),
    }


# ── 헬퍼 ───────────────────────────────────────────────────────────────────────
def to_num(val):
    if val is None or (isinstance(val, float) and str(val) == 'nan'):
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).replace(',', '').strip()
    if s in ('', 'nan'):
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return val


def norm_date(val):
    if val is None:
        return None
    s = str(val)[:10].replace('-', '.')
    if not re.match(r'^\d{4}\.\d{2}\.\d{2}$', s):
        return None
    return s


def get_cell_val(cell):
    """수식 셀(="2026.06.17" 형태)도 실제 값으로 반환"""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str) and v.startswith('="') and v.endswith('"'):
        return v[2:-1]
    return v


def get_max_row(ws):
    return ws.max_row or 1


def copy_row_format(ws, src_row, dst_row):
    """윗 행 서식을 새 행에 복사"""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font      = copy(src.font)
            dst.fill      = copy(src.fill)
            dst.border    = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def read_col(ws, col, start_row, end_row):
    if end_row < start_row:
        return []
    return [get_cell_val(ws.cell(r, col)) for r in range(start_row, end_row + 1)]


def read_block(ws, start_row, start_col, end_row, end_col):
    if end_row < start_row:
        return []
    result = []
    for r in range(start_row, end_row + 1):
        result.append([ws.cell(r, c).value for c in range(start_col, end_col + 1)])
    return result


def write_row(ws, row, start_col, values):
    for i, v in enumerate(values):
        ws.cell(row, start_col + i).value = v


# ── 파서 ───────────────────────────────────────────────────────────────────────
def parse_musinsa_xls(fpath):
    with open(fpath, 'r', encoding='utf-8-sig') as f:
        content = f.read()
    content = re.sub(r' xmlns[^=]*="[^"]*"', '', content)
    content = re.sub(r'<\?[^>]+\?>', '', content)
    content = re.sub(r'ss:', '', content)
    root = ET.fromstring(content)
    result = []
    for ws_elem in root.findall('.//Worksheet'):
        table = ws_elem.find('.//Table')
        if table is None:
            continue
        for i, row_elem in enumerate(table.findall('Row')):
            if i < 1:
                continue
            cells = [c.findtext('Data', '') for c in row_elem.findall('Cell')]
            if not cells or not cells[0]:
                continue
            result.append((cells[0], [to_num(v) for v in cells[1:]]))
    return result


def parse_global_orders(fpath):
    dfs = pd.read_html(fpath, encoding='utf-8', header=0,
                       converters={1: str, 2: str})
    if not dfs:
        raise ValueError("테이블 없음: " + fpath)
    df = dfs[0].copy()
    df = df.drop_duplicates(subset=[df.columns[2]], keep='last')
    return df


def parse_29cm_report(fpath):
    today = _date.today().strftime('%Y-%m-%d')
    df = pd.read_excel(fpath, header=0)
    result = []
    for _, row in df.iterrows():
        date_raw = str(row.iloc[0])[:10]
        if not date_raw.startswith('2026'):
            continue
        if date_raw > today:
            continue
        result.append((date_raw, [to_num(v) for v in row.iloc[1:12]]))
    return result


# ── 유령 행 정리 ────────────────────────────────────────────────────────────────
def clean_phantom_rows(ws, col_a, col_b, data_row_start=2, label=''):
    """col_a, col_b 두 날짜 컬럼이 모두 빈(None) 행을 삭제 (insert_rows 잔재 정리)."""
    max_r = get_max_row(ws)
    if max_r < data_row_start:
        return

    vals_a = read_col(ws, col_a, data_row_start, max_r)
    vals_b = read_col(ws, col_b, data_row_start, max_r)

    to_delete = [
        data_row_start + i for i in range(len(vals_a))
        if vals_a[i] is None and (i >= len(vals_b) or vals_b[i] is None)
    ]
    if not to_delete:
        return

    to_delete.sort(reverse=True)
    i = 0
    while i < len(to_delete):
        start = to_delete[i]
        end = start
        while i + 1 < len(to_delete) and to_delete[i + 1] == end - 1:
            end = to_delete[i + 1]
            i += 1
        ws.delete_rows(end, start - end + 1)
        i += 1
    print(f"  [{label}] 유령 행 {len(to_delete)}개 정리 완료")


# ── upsert 함수 ─────────────────────────────────────────────────────────────────
def upsert_special_rows(ws, rows, date_col, data_col_start, num_data,
                        avg_row=2, data_row_start=3, label=''):
    avg_vals = sum_vals = None
    for date_str, vals in rows:
        s = str(date_str).strip()
        if s == '평균':
            avg_vals = vals
        elif s == '합계':
            sum_vals = vals

    if avg_vals is not None:
        write_row(ws, avg_row, data_col_start,
                  [avg_vals[i] if i < len(avg_vals) else None for i in range(num_data)])
        print(f"  [{label}] 평균 행(row {avg_row}) 업데이트")
    else:
        print(f"  [{label}] 평균 행 데이터 없음")

    if sum_vals is not None:
        max_r = get_max_row(ws)
        date_col_vals = read_col(ws, date_col, data_row_start, max_r)
        sum_row = next(
            (data_row_start + i for i, v in enumerate(date_col_vals)
             if v is not None and str(v).strip() == '합계'),
            None
        )
        if sum_row:
            write_row(ws, sum_row, data_col_start,
                      [sum_vals[i] if i < len(sum_vals) else None for i in range(num_data)])
            print(f"  [{label}] 합계 행(row {sum_row}) 업데이트")
        else:
            print(f"  [{label}] 합계 행 없음")
    else:
        print(f"  [{label}] 합계 행 데이터 없음")


def upsert_by_date(ws, rows, date_col, data_col_start, num_data,
                   data_row_start=3, a_formula_tpl=None,
                   stop_keyword='합계', pair_date_col=None,
                   extra_formulas=None, label=''):
    max_r = get_max_row(ws)

    date_vals = read_col(ws, date_col, data_row_start, max_r)
    date_to_row = {}
    sum_row = None
    last_data_row = data_row_start - 1

    for i, val in enumerate(date_vals):
        r = data_row_start + i
        if val is None:
            continue
        val_s = str(val).strip()
        if stop_keyword and val_s == stop_keyword:
            sum_row = r
            break
        d = norm_date(val_s)
        if d:
            date_to_row[d] = r
            last_data_row = r

    row_cache = {}
    if date_to_row:
        min_r = min(date_to_row.values())
        max_d_r = max(date_to_row.values())
        block = read_block(ws, min_r, data_col_start, max_d_r, data_col_start + num_data - 1)
        for i, row_vals in enumerate(block):
            row_cache[min_r + i] = row_vals if isinstance(row_vals, list) else [row_vals]

    pair_date_to_row = {}
    if pair_date_col:
        pair_vals = read_col(ws, pair_date_col, data_row_start, max_r)
        for i, val in enumerate(pair_vals):
            if val is None:
                continue
            d = norm_date(str(val).strip())
            if d:
                pair_date_to_row[d] = data_row_start + i

    updated = added = skipped = 0

    for date_str, vals in rows:
        if not str(date_str).startswith('2026'):
            continue
        d = norm_date(date_str)

        if d in date_to_row:
            r = date_to_row[d]
            new_vals = [vals[i] if i < len(vals) else None for i in range(num_data)]
            existing = row_cache.get(r, [None] * num_data)
            if existing[:num_data] != new_vals:
                write_row(ws, r, data_col_start, new_vals)
                updated += 1
            else:
                skipped += 1

        else:
            if pair_date_col and d in pair_date_to_row:
                new_r = pair_date_to_row[d]
            elif sum_row is not None:
                ws.insert_rows(sum_row)
                new_r = sum_row
                sum_row += 1
                pair_date_to_row = {k: (v + 1 if v >= new_r else v)
                                    for k, v in pair_date_to_row.items()}
                copy_row_format(ws, new_r - 1, new_r)
            else:
                last_data_row += 1
                new_r = last_data_row
                copy_row_format(ws, new_r - 1, new_r)

            if a_formula_tpl:
                ws.cell(new_r, 1).value = a_formula_tpl.format(r=new_r)
            if extra_formulas:
                for ecol, etpl in extra_formulas.items():
                    ws.cell(new_r, ecol).value = etpl.format(r=new_r)

            ws.cell(new_r, date_col).value = f'="{date_str}"'
            write_row(ws, new_r, data_col_start,
                      [vals[i] if i < len(vals) else None for i in range(num_data)])
            date_to_row[d] = new_r
            last_data_row = max(last_data_row, new_r)
            added += 1

    print(f"  [{label}] 업데이트 {updated}건 | 신규 {added}건 | 동일 {skipped}건")
    return updated, added, skipped


def upsert_global_orders(ws, df, key_sheet_col=15,
                         data_sheet_start_col=13, label='글로벌'):
    KEY_DF_COL = 2
    num_cols = len(df.columns)
    max_r = get_max_row(ws)

    key_vals = read_col(ws, key_sheet_col, 3, max_r)
    key_to_row = {}
    last_row = 2
    for i, kv in enumerate(key_vals):
        if kv is not None:
            key_to_row[str(kv).strip()] = 3 + i
            last_row = 3 + i

    row_cache = {}
    if last_row >= 3:
        block = read_block(ws, 3, data_sheet_start_col,
                           last_row, data_sheet_start_col + num_cols - 1)
        for i, rv in enumerate(block):
            row_cache[3 + i] = rv if isinstance(rv, list) else [rv]

    updated = added = skipped = 0

    for _, src_row in df.iterrows():
        key_val = str(src_row.iloc[KEY_DF_COL]).strip()
        if key_val in ('', 'nan', 'None'):
            continue
        new_vals = [None if str(v) in ('nan', 'None', '') else v for v in list(src_row)]

        if key_val in key_to_row:
            r = key_to_row[key_val]
            existing = row_cache.get(r, [None] * num_cols)
            if [str(a) for a in existing] != [str(b) for b in new_vals]:
                write_row(ws, r, data_sheet_start_col, new_vals)
                updated += 1
            else:
                skipped += 1
        else:
            last_row += 1
            copy_row_format(ws, last_row - 1, last_row)
            write_row(ws, last_row, data_sheet_start_col, new_vals)
            key_to_row[key_val] = last_row
            added += 1

    print(f"  [{label}] 업데이트 {updated}건 | 신규 {added}건 | 동일 {skipped}건")
    return updated, added, skipped


# ── 메인 ───────────────────────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  판매 통계 자동 업데이트 시작")
    print("=" * 55)

    files = find_source_files()
    print("\n[소스 파일 탐지]")
    for k, v in files.items():
        print(f"  {k}: {os.path.basename(v) if v else '없음'}")

    if not os.path.exists(STATS_FILE):
        print(f"\n오류: 메인 파일 없음 → {STATS_FILE}")
        sys.exit(1)

    print(f"\n[백업] 판매 통계.xlsx → 판매 통계 백업.xlsx")
    shutil.copy2(STATS_FILE, BACKUP_FILE)
    print("  백업 완료")

    import time
    t0 = time.time()

    print("  [openpyxl] 파일 로드 중...")
    wb = openpyxl.load_workbook(STATS_FILE)
    print("  [openpyxl] 로드 완료")

    results = {}

    # 1. 무신사
    print("\n[1] 무신사 시트 업데이트")
    ws_m = wb['무신사']
    clean_phantom_rows(ws_m, col_a=2, col_b=20, data_row_start=3, label='무신사')
    if files["mushinsa"]:
        rows_c = parse_musinsa_xls(files["mushinsa"])
        r = upsert_by_date(
            ws_m, rows_c, date_col=2, data_col_start=3, num_data=14,
            data_row_start=3,
            a_formula_tpl='=IFERROR(VALUE(SUBSTITUTE(B{r},".","-")),"?")',
            stop_keyword='합계', label='커넥트킨록')
        upsert_special_rows(ws_m, rows_c, date_col=2, data_col_start=3,
                            num_data=14, avg_row=2, data_row_start=3,
                            label='커넥트킨록 특수행')
        results['무신사_커넥트'] = r
    else:
        print("  [커넥트킨록] 파일 없음, 스킵")

    if files["mushinsa2"]:
        rows_e = parse_musinsa_xls(files["mushinsa2"])
        r = upsert_by_date(
            ws_m, rows_e, date_col=20, data_col_start=21, num_data=14,
            data_row_start=3, stop_keyword='합계', pair_date_col=2,
            extra_formulas={19: '=IFERROR(VALUE(SUBSTITUTE(T{r}, ".", "-")),"")'},
            label='에든버러클럽')
        upsert_special_rows(ws_m, rows_e, date_col=20, data_col_start=21,
                            num_data=14, avg_row=2, data_row_start=3,
                            label='에든버러클럽 특수행')
        results['무신사_에든버러'] = r
    else:
        print("  [에든버러클럽] 파일 없음, 스킵")

    # 2. 글로벌
    print("\n[2] 글로벌 시트 업데이트")
    if files["glob"]:
        df_global = parse_global_orders(files["glob"])
        r = upsert_global_orders(
            wb['글로벌'], df_global,
            key_sheet_col=15, data_sheet_start_col=13)
        results['글로벌'] = r
        del df_global
        gc.collect()
    else:
        print("  [주문내역] 파일 없음, 스킵")

    # 3. 29CM
    print("\n[3] 29CM 시트 업데이트")
    ws_cm = wb['29CM']
    clean_phantom_rows(ws_cm, col_a=2, col_b=17, data_row_start=2, label='29CM')
    if files["cm1"]:
        rows_cm_c = parse_29cm_report(files["cm1"])
        r = upsert_by_date(
            ws_cm, rows_cm_c, date_col=2, data_col_start=3, num_data=11,
            data_row_start=2, a_formula_tpl='=VALUE(B{r})',
            stop_keyword=None, label='29CM 커넥트킨록')
        results['29CM_커넥트'] = r
    else:
        print("  [29CM 커넥트] 파일 없음, 스킵")

    if files["cm2"]:
        rows_cm_e = parse_29cm_report(files["cm2"])
        r = upsert_by_date(
            ws_cm, rows_cm_e, date_col=17, data_col_start=18, num_data=11,
            data_row_start=2, stop_keyword=None, pair_date_col=2,
            extra_formulas={16: '=VALUE(Q{r})'},
            label='29CM 에든버러클럽')
        results['29CM_에든버러'] = r
    else:
        print("  [29CM 에든버러] 파일 없음, 스킵")

    elapsed = time.time() - t0
    print(f"\n[저장] 판매 통계.xlsx 저장 중...")
    wb.save(STATS_FILE)
    print(f"  저장 완료 ({elapsed:.1f}초)")

    print("\n" + "=" * 55)
    print("  업데이트 완료 요약")
    print("=" * 55)
    labels = {
        '무신사_커넥트':   '무신사   커넥트킨록',
        '무신사_에든버러': '무신사   에든버러클럽',
        '글로벌':         '글로벌   주문내역',
        '29CM_커넥트':    '29CM     커넥트킨록',
        '29CM_에든버러':  '29CM     에든버러클럽',
    }
    for key, label in labels.items():
        if key in results:
            u, a, s = results[key]
            parts = []
            if a: parts.append(f"신규 {a}건")
            if u: parts.append(f"수정 {u}건")
            if s: parts.append(f"동일 {s}건")
            print(f"  {label}: {' | '.join(parts) if parts else '변경 없음'}")
        else:
            print(f"  {label}: 파일 없음 (스킵)")

    print(f"\n  저장 완료 → 판매 통계.xlsx")
    print(f"  백업 완료 → 판매 통계 백업.xlsx")
    print("=" * 55)


if __name__ == '__main__':
    main()
